import datetime
import os
from openpyxl import load_workbook
from src.utility.excel_handler import *
import win32com.client as client
import pythoncom
from pathlib import Path


def create_takeoff_file(src: Path, dest: Path, proj_data: dict):
    """
    Copy the takeoff template file, add information to it and save it as xlsm.

    :param src: source path of the template file
    :param dest: destination to save the xlsm file in
    :param proj_data: object containing project information
    """
    today_date = datetime.datetime.today().strftime('%d-%m-%Y')
    key_value_pairs = [
        ['XXXX', proj_data['id']],
        ['TODAYSDATE', today_date],
        ['CLIENT', proj_data['client']],
        ['PROJECT NAME', proj_data['name']],
        ['PROJECT TYPE', proj_data['type']],
        ['PROPOSAL-URL', proj_data['url']]
    ]
    takeoff_wb = load_workbook(src, read_only=False, keep_vba=True)
    takeoff_ws = takeoff_wb.worksheets[0]
    change_cells_with_values(takeoff_ws, key_value_pairs)
    takeoff_wb.save(dest)
    save_as_xlsm(dest, dest)
    os.remove(dest)


def create_checklist_file(src: Path, dest: Path, proj_data: dict):
    """
    Copy the checklist template file, add information to it and save it as xlsm.

    :param src: source path of the template file
    :param dest: destination to save the xlsm file in
    :param proj_data: object containing project information
    """
    checklist_wb = load_workbook(src, read_only=False, keep_vba=True)
    checklist_ws = checklist_wb.worksheets[0]
    change_adjacent_cell(checklist_ws, 'PROJECT#', proj_data['id'])
    checklist_wb.save(dest)
    save_as_xlsm(dest, dest)
    os.remove(dest)


def create_config_file(src: Path, dest: Path, proj_data: dict):
    """
    Copy the configurator template file, add information to it and save it as xlsm.

    :param src: source path of the template file
    :param dest: destination to save the xlsm file in
    :param proj_data: object containing project information
    """
    key_value_pairs = [
        ['Project Number:', proj_data['id']],
        ['Project Name:', proj_data['name']]
    ]
    config_wb = load_workbook(src, read_only=False, keep_vba=True)
    config_ws = config_wb.worksheets[0]
    change_adjacent_cells_with_values(config_ws, key_value_pairs)
    config_wb.save(dest)
    save_as_xlsm(dest, dest)
    os.remove(dest)


def update_quote_log(path, proj_data):
    """
    Add a row entry for the project in the Quote Log.

    :param path: full path of the Quote Log file
    :param proj_data: object containing project information
    """
    ql_wb = load_workbook(path, read_only=False, keep_vba=True)
    ql_ws = ql_wb.worksheets[0]
    last_row = get_last_empty_row(ql_ws, 'B')

    today_date = datetime.datetime.today().strftime('%d-%b-%y')
    row_data = [
        today_date,
        proj_data['client'],
        proj_data['path'],
        proj_data['item'],
        proj_data['id'],
        'TBD',
        proj_data['rep']
    ]
    fill_row_with_values(ql_ws, last_row, row_data)
    ql_wb.save(path)


def save_as_xlsm(src: Path, dest: Path):
    """
    Save the given xltm file as xlsm.

    :param src: source path of the template file
    :param dest: destination to save the xlsm file in
    """
    app, command = 'Excel.Application', pythoncom.CoInitialize
    excel = client.gencache.EnsureDispatch(app, command())
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(str(src))
    filename = str(dest.parent / (dest.stem + '.xlsm'))
    wb.SaveAs(Filename=filename, FileFormat=52, CreateBackup=False)
    wb.Close()
